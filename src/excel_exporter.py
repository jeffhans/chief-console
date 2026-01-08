"""
Excel Exporter for CP4I Mission Console
Generates customer-friendly Excel spreadsheets for licensing and sizing collaboration
"""

from typing import Dict, Any, List, Optional
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Warning: openpyxl not installed. Run: pip3 install openpyxl")
    raise


class ExcelExporter:
    """Export Mission Console data to Excel format for customer collaboration"""

    def __init__(self, snapshot: Dict[str, Any], resource_summary: Optional[Dict[str, Any]] = None):
        """
        Initialize Excel exporter

        Args:
            snapshot: Full snapshot data from collector
            resource_summary: Resource categorization summary (optional)
        """
        self.snapshot = snapshot
        self.resource_summary = resource_summary or {}
        self.wb = Workbook()
        self.cluster = self.snapshot.get('cluster', {}) or {}
        self.cluster_name = (
            self.cluster.get('alias')
            or self.cluster.get('display_name')
            or self.cluster.get('id')
            or self.snapshot.get('metadata', {}).get('cluster_display')
            or 'Unknown'
        )
        self.cluster_id = (
            self.cluster.get('id')
            or self.snapshot.get('metadata', {}).get('cluster_id')
            or 'Unknown'
        )
        self.cluster_version = self.cluster.get('version', 'Unknown')
        self.cluster_alias_id = (
            self.cluster.get('alias_id')
            or self.snapshot.get('metadata', {}).get('cluster_alias_id')
            or ''
        )
        self.last_updated = self.snapshot.get('metadata', {}).get('collection_timestamp', '')

        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])

        # Styling
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.subheader_font = Font(color="FFFFFF", bold=True, size=10)
        self.highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export(self, output_path: str) -> str:
        """
        Generate Excel file with all sheets

        Args:
            output_path: Path to save Excel file

        Returns:
            Path to created file
        """
        print("\n📊 Generating Excel export...")

        # Create all sheets
        self._create_executive_summary()
        self._create_licensing_sheet()
        self._create_infrastructure_sheet()
        self._create_workload_sheet()
        self._create_detailed_pods_sheet()

        # Save workbook
        self.wb.save(output_path)
        print(f"✓ Excel file created: {output_path}")

        return output_path

    def _create_executive_summary(self):
        """Sheet 1: Executive Summary - High-level overview"""
        ws = self.wb.create_sheet("Executive Summary", 0)

        # Title
        ws['A1'] = "CP4I Mission Console - Executive Summary"
        ws['A1'].font = Font(size=14, bold=True, color="1F4E78")
        ws.merge_cells('A1:D1')

        # Report metadata
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Name: {self.cluster_name}"
        ws['A4'] = f"ID: {self.cluster_alias_id}"
        ws['A5'] = f"OpenShift {self.cluster_version}"
        ws['A6'] = f"Last updated: {self._format_timestamp(self.last_updated)}" if self.last_updated else "Last updated: Unknown"
        ws['A7'] = f"Cluster ID: {self.cluster_id}"
        ws['A3'].font = Font(bold=True)
        ws['A4'].font = Font(bold=True)
        ws['A5'].font = Font(bold=True)
        ws['A6'].font = Font(bold=True)
        ws['A7'].font = Font(bold=True)

        row = 9

        # Cluster Overview Section
        ws[f'A{row}'] = "CLUSTER OVERVIEW"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        nodes = self.snapshot.get('nodes', [])

        # Cluster metrics
        total_cpu = sum(int(n.get('capacity', {}).get('cpu', 0)) for n in nodes)
        total_memory_ki = sum(self._parse_memory_ki(n.get('capacity', {}).get('memory', '0Ki')) for n in nodes)
        total_memory_gib = total_memory_ki / (1024 * 1024)

        metrics = [
            ("OpenShift Version", self.cluster_version),
            ("Total Nodes", len(nodes)),
            ("Total CPU Cores", total_cpu),
            ("Total Memory", f"{total_memory_gib:.1f} GiB"),
        ]

        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        row += 1

        # Licensing Section
        ws[f'A{row}'] = "CP4I LICENSING"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        licensing = self.resource_summary.get('licensing', {})
        breakdown = licensing.get('breakdown', {})
        total_vpc = licensing.get('total_vpc', 0)

        ws[f'A{row}'] = "Total VPC Consumption"
        ws[f'B{row}'] = f"{total_vpc:.2f}"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].font = Font(bold=True, size=12, color="C00000")
        ws[f'B{row}'].fill = self.highlight_fill
        row += 1

        ws[f'A{row}'] = "CP4I Licensed Pods"
        ws[f'B{row}'] = breakdown.get('cp4i_licensed', 0)
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = "Platform Pods"
        ws[f'B{row}'] = breakdown.get('openshift_platform', 0)
        row += 1

        ws[f'A{row}'] = "Free/OSS Pods"
        ws[f'B{row}'] = breakdown.get('free', 0)
        row += 1

        row += 1

        # Workload Section
        ws[f'A{row}'] = "WORKLOAD SUMMARY"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        workloads = self.resource_summary.get('workloads', {})
        business_count = workloads.get('business_workloads', 0)
        infra_count = workloads.get('infrastructure', 0)

        ws[f'A{row}'] = "Business Workloads"
        ws[f'B{row}'] = business_count
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = "Infrastructure Pods"
        ws[f'B{row}'] = infra_count
        row += 1

        row += 1

        # Resource Utilization
        ws[f'A{row}'] = "RESOURCE REQUESTS"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        resources = self.resource_summary.get('resources', {})
        cpu_requests = resources.get('total_cpu_requests_cores', 0)
        memory_requests_gb = resources.get('total_memory_requests_gb', 0)

        ws[f'A{row}'] = "Total CPU Requests"
        ws[f'B{row}'] = f"{cpu_requests:.2f} cores"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = "Total Memory Requests"
        ws[f'B{row}'] = f"{memory_requests_gb:.2f} GB"
        row += 1

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25

    def _create_licensing_sheet(self):
        """Sheet 2: Licensing Analysis - CP4I VPC consumption by capability"""
        ws = self.wb.create_sheet("Licensing Analysis")

        # Header
        ws['A1'] = "CP4I Licensing Analysis"
        ws['A1'].font = Font(size=14, bold=True, color="1F4E78")
        ws.merge_cells('A1:F1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Summary box
        row = 4
        ws[f'A{row}'] = "LICENSING SUMMARY"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        licensing = self.resource_summary.get('licensing', {})
        total_vpc = licensing.get('total_vpc', 0)

        ws[f'A{row}'] = "Total VPC Consumption:"
        ws[f'B{row}'] = f"{total_vpc:.2f}"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].font = Font(bold=True, size=12, color="C00000")
        ws[f'B{row}'].fill = self.highlight_fill
        row += 2

        # Get CP4I licensed pods
        categorized = self.resource_summary.get('categorized_pods', [])
        cp4i_pods = [p for p in categorized if p.get('licensing') == 'cp4i_licensed']

        # Group by capability (namespace prefix)
        capability_groups = {}
        for pod in cp4i_pods:
            namespace = pod.get('namespace', '')
            # Group by namespace
            if namespace not in capability_groups:
                capability_groups[namespace] = []
            capability_groups[namespace].append(pod)

        # Table header
        ws[f'A{row}'] = "CP4I COMPONENTS BY CAPABILITY"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        headers = ["Namespace/Capability", "Pod Count", "CPU Requests (cores)", "Memory Requests (GB)", "VPC Contribution", "Notes"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        row += 1

        # Capability rows
        for capability, pods in sorted(capability_groups.items()):
            pod_count = len(pods)
            cpu_total = sum(p.get('resources', {}).get('cpu_requests', 0) for p in pods)
            memory_total = sum(p.get('resources', {}).get('memory_requests', 0) for p in pods)
            memory_gb = memory_total / (1024**3)
            vpc_total = sum(p.get('vpc', 0) for p in pods)

            # Identify capability type
            capability_name = capability
            notes = ""
            if 'ibm-eventstreams' in capability or 'es-demo' in capability:
                notes = "Event Streams"
            elif 'ibm-ep' in capability or 'ep-demo' in capability:
                notes = "Event Processing"
            elif 'ace' in capability or 'appconnect' in capability:
                notes = "App Connect"
            elif 'mq' in capability:
                notes = "MQ"
            elif 'apic' in capability:
                notes = "API Connect"
            elif 'aspera' in capability:
                notes = "Aspera"
            elif 'assetrepo' in capability:
                notes = "Asset Repository"

            ws.cell(row=row, column=1, value=capability_name).border = self.border
            ws.cell(row=row, column=2, value=pod_count).border = self.border
            ws.cell(row=row, column=3, value=f"{cpu_total:.2f}").border = self.border
            ws.cell(row=row, column=4, value=f"{memory_gb:.2f}").border = self.border
            ws.cell(row=row, column=5, value=f"{vpc_total:.2f}").border = self.border
            ws.cell(row=row, column=5).font = Font(bold=True, color="C00000")
            ws.cell(row=row, column=6, value=notes).border = self.border
            row += 1

        # Totals row
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=len(cp4i_pods)).font = Font(bold=True)
        total_cpu = sum(p.get('resources', {}).get('cpu_requests', 0) for p in cp4i_pods)
        total_memory = sum(p.get('resources', {}).get('memory_requests', 0) for p in cp4i_pods)
        ws.cell(row=row, column=3, value=f"{total_cpu:.2f}").font = Font(bold=True)
        ws.cell(row=row, column=4, value=f"{total_memory / (1024**3):.2f}").font = Font(bold=True)
        ws.cell(row=row, column=5, value=f"{total_vpc:.2f}").font = Font(bold=True, color="C00000")

        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = self.highlight_fill
            ws.cell(row=row, column=col).border = self.border

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 20

    def _create_infrastructure_sheet(self):
        """Sheet 3: Infrastructure Sizing - Node capacity and utilization"""
        ws = self.wb.create_sheet("Infrastructure Sizing")

        # Header
        ws['A1'] = "Infrastructure Sizing"
        ws['A1'].font = Font(size=14, bold=True, color="1F4E78")
        ws.merge_cells('A1:F1')

        ws['A2'] = f"Name: {self.cluster_name}"
        ws['A3'] = f"ID: {self.cluster_alias_id}"
        ws['A4'] = f"OpenShift {self.cluster_version}"
        ws['A5'] = f"Last updated: {self._format_timestamp(self.last_updated)}" if self.last_updated else "Last updated: Unknown"
        ws['A6'] = f"Cluster ID: {self.cluster_id}"
        ws['A2'].font = Font(bold=True)
        ws['A3'].font = Font(bold=True)
        ws['A4'].font = Font(bold=True)
        ws['A5'].font = Font(bold=True)
        ws['A6'].font = Font(bold=True)

        row = 8

        # Cluster capacity summary
        nodes = self.snapshot.get('nodes', [])
        total_cpu = sum(int(n.get('capacity', {}).get('cpu', 0)) for n in nodes)
        total_memory_ki = sum(self._parse_memory_ki(n.get('capacity', {}).get('memory', '0Ki')) for n in nodes)
        total_memory_gib = total_memory_ki / (1024 * 1024)

        ws[f'A{row}'] = "CLUSTER CAPACITY"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        ws[f'A{row}'] = "Total Nodes"
        ws[f'B{row}'] = len(nodes)
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = "Total CPU Cores"
        ws[f'B{row}'] = total_cpu
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = "Total Memory"
        ws[f'B{row}'] = f"{total_memory_gib:.1f} GiB"
        ws[f'A{row}'].font = Font(bold=True)
        row += 2

        # Resource requests
        resources = self.resource_summary.get('resources', {})
        cpu_requests = resources.get('total_cpu_requests_cores', 0)
        memory_requests_gb = resources.get('total_memory_requests_gb', 0)

        ws[f'A{row}'] = "RESOURCE REQUESTS (Allocated)"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        ws[f'A{row}'] = "CPU Requests"
        ws[f'B{row}'] = f"{cpu_requests:.2f} cores"
        ws[f'C{row}'] = f"{(cpu_requests/total_cpu*100) if total_cpu > 0 else 0:.1f}% of capacity"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = "Memory Requests"
        ws[f'B{row}'] = f"{memory_requests_gb:.2f} GB"
        ws[f'C{row}'] = f"{(memory_requests_gb/total_memory_gib*100) if total_memory_gib > 0 else 0:.1f}% of capacity"
        ws[f'A{row}'].font = Font(bold=True)
        row += 2

        # Node details table
        ws[f'A{row}'] = "NODE DETAILS"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        headers = ["Node Name", "Status", "CPU", "Memory", "Pods", "Version"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        row += 1

        for node in nodes:
            name = node.get('name', 'Unknown')
            status = node.get('status', 'Unknown')
            cpu = node.get('capacity', {}).get('cpu', 'N/A')
            memory_ki = self._parse_memory_ki(node.get('capacity', {}).get('memory', '0Ki'))
            memory_gib = memory_ki / (1024 * 1024)
            pods = node.get('capacity', {}).get('pods', 'N/A')
            version = node.get('version', 'Unknown')

            ws.cell(row=row, column=1, value=name).border = self.border
            ws.cell(row=row, column=2, value=status).border = self.border
            ws.cell(row=row, column=3, value=cpu).border = self.border
            ws.cell(row=row, column=4, value=f"{memory_gib:.1f} GiB").border = self.border
            ws.cell(row=row, column=5, value=pods).border = self.border
            ws.cell(row=row, column=6, value=version).border = self.border
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 30

    def _create_workload_sheet(self):
        """Sheet 4: Workload Inventory - Business applications"""
        ws = self.wb.create_sheet("Workload Inventory")

        # Header
        ws['A1'] = "Workload Inventory"
        ws['A1'].font = Font(size=14, bold=True, color="1F4E78")
        ws.merge_cells('A1:G1')

        ws['A2'] = "Business workloads only (excludes infrastructure components)"

        row = 4

        # Get business workloads
        categorized = self.resource_summary.get('categorized_pods', [])
        business_pods = [p for p in categorized if p.get('is_workload', False)]

        ws[f'A{row}'] = f"BUSINESS WORKLOADS ({len(business_pods)} total)"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:G{row}')
        row += 1

        if len(business_pods) == 0:
            ws[f'A{row}'] = "No business workloads detected"
            ws[f'A{row}'].font = Font(italic=True, color="666666")
            row += 1
            ws[f'A{row}'] = "Business workloads are applications like healthcare-fhir-enricher, order-processing, etc."
            ws[f'A{row}'].font = Font(italic=True, size=9, color="666666")
        else:
            # Table header
            headers = ["Workload Name", "Namespace", "Status", "CPU Requests", "Memory Requests", "Criticality", "VPC"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            row += 1

            # Workload rows
            for pod in sorted(business_pods, key=lambda x: x.get('namespace', '')):
                name = pod.get('name', 'Unknown')
                namespace = pod.get('namespace', 'Unknown')
                phase = pod.get('phase', 'Unknown')
                cpu = pod.get('resources', {}).get('cpu_requests', 0)
                memory = pod.get('resources', {}).get('memory_requests', 0)
                memory_gb = memory / (1024**3)
                criticality = pod.get('criticality', 'optional')
                vpc = pod.get('vpc', 0)

                ws.cell(row=row, column=1, value=name).border = self.border
                ws.cell(row=row, column=2, value=namespace).border = self.border
                ws.cell(row=row, column=3, value=phase).border = self.border
                ws.cell(row=row, column=4, value=f"{cpu:.2f}").border = self.border
                ws.cell(row=row, column=5, value=f"{memory_gb:.2f} GB").border = self.border
                ws.cell(row=row, column=6, value=criticality).border = self.border
                ws.cell(row=row, column=7, value=f"{vpc:.2f}").border = self.border
                row += 1

        # Column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12

    def _create_detailed_pods_sheet(self):
        """Sheet 5: Detailed Pod List - Complete resource breakdown"""
        ws = self.wb.create_sheet("Detailed Pod List")

        # Header
        ws['A1'] = "Detailed Pod List"
        ws['A1'].font = Font(size=14, bold=True, color="1F4E78")
        ws.merge_cells('A1:H1')

        ws['A2'] = "Complete inventory of all pods with licensing and resource details"

        row = 4

        # Table header
        headers = ["Pod Name", "Namespace", "Status", "Licensing", "Criticality", "CPU Requests", "Memory Requests", "VPC"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        row += 1

        # Pod rows
        categorized = self.resource_summary.get('categorized_pods', [])
        for pod in sorted(categorized, key=lambda x: (x.get('namespace', ''), x.get('name', ''))):
            name = pod.get('name', 'Unknown')
            namespace = pod.get('namespace', 'Unknown')
            phase = pod.get('phase', 'Unknown')
            licensing = pod.get('licensing', 'unknown')
            criticality = pod.get('criticality', 'optional')
            cpu = pod.get('resources', {}).get('cpu_requests', 0)
            memory = pod.get('resources', {}).get('memory_requests', 0)
            memory_gb = memory / (1024**3)
            vpc = pod.get('vpc', 0)

            ws.cell(row=row, column=1, value=name).border = self.border
            ws.cell(row=row, column=2, value=namespace).border = self.border
            ws.cell(row=row, column=3, value=phase).border = self.border

            # Color code licensing
            lic_cell = ws.cell(row=row, column=4, value=licensing)
            lic_cell.border = self.border
            if licensing == 'cp4i_licensed':
                lic_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                lic_cell.font = Font(color="C00000", bold=True)
            elif licensing == 'openshift_platform':
                lic_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            ws.cell(row=row, column=5, value=criticality).border = self.border
            ws.cell(row=row, column=6, value=f"{cpu:.2f}").border = self.border
            ws.cell(row=row, column=7, value=f"{memory_gb:.2f} GB").border = self.border

            # Highlight VPC contribution
            vpc_cell = ws.cell(row=row, column=8, value=f"{vpc:.2f}")
            vpc_cell.border = self.border
            if vpc > 0:
                vpc_cell.fill = self.highlight_fill
                vpc_cell.font = Font(bold=True, color="C00000")

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 12

    def _parse_memory_ki(self, memory_str: str) -> int:
        """Parse memory string to Ki (kibibytes)"""
        if not memory_str or memory_str == 'N/A':
            return 0

        try:
            if memory_str.endswith('Ki'):
                return int(memory_str[:-2])
            elif memory_str.endswith('Mi'):
                return int(memory_str[:-2]) * 1024
            elif memory_str.endswith('Gi'):
                return int(memory_str[:-2]) * 1024 * 1024
            elif memory_str.endswith('Ti'):
                return int(memory_str[:-2]) * 1024 * 1024 * 1024
            else:
                return int(memory_str)
        except (ValueError, AttributeError):
            return 0

    def _format_timestamp(self, ts: str) -> str:
        """Format ISO timestamp for display"""
        if not ts:
            return ""
        try:
            dt = datetime.fromisoformat(ts.replace('Z', '+00:00'))
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            return ts
